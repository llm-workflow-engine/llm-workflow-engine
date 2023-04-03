import os
import csv
import json
import yaml
import configparser
import openpyxl
import xml.etree.ElementTree as ET
from prompt_toolkit.completion import PathCompleter

from langchain.agents import create_json_agent
from langchain.agents.agent_toolkits import JsonToolkit
from langchain.tools.json.tool import JsonSpec

from chatgpt_wrapper.core.plugin import Plugin
from chatgpt_wrapper.core.config import Config
from chatgpt_wrapper.core.logger import Logger
import chatgpt_wrapper.core.util as util

class DataLoader:
    def __init__(self, config):
        self.config = config or Config()
        self.log = Logger(self.__class__.__name__, self.config)
        self.data = None

    def load(self, filepath):
        self.filepath = filepath
        self.file_extension = os.path.splitext(self.filepath)[1].lower()
        if self.file_extension == '.csv':
            return self.load_csv()
        elif self.file_extension in ['.xlsx']:
            return self.load_excel()
        elif self.file_extension == '.json':
            return self.load_json()
        elif self.file_extension == '.jsonl':
            return self.load_jsonl()
        elif self.file_extension == '.yaml':
            return self.load_yaml()
        elif self.file_extension == '.ini':
            return self.load_ini()
        elif self.file_extension == '.geojson':
            return self.load_geojson()
        elif self.file_extension == '.xml':
            return self.load_xml()
        else:
            return False, None, "Unsupported file format."

    def load_csv(self):
        try:
            with open(self.filepath, newline='', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                data = [row for row in reader]
            return True, data, "CSV file loaded successfully."
        except Exception as e:
            return False, None, f"Error loading CSV file: {e}"

    def load_excel(self):
        try:
            wb = openpyxl.load_workbook(self.filepath)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            data = [
                {header[i]: cell.value for i, cell in enumerate(row)}
                for row in ws.iter_rows(min_row=2)
            ]
            return True, data, "Excel file loaded successfully."
        except Exception as e:
            return False, None, f"Error loading Excel file: {e}"

    def load_json(self):
        try:
            with open(self.filepath, 'r', encoding='utf-8') as jsonfile:
                data = json.load(jsonfile)
            return True, data, "JSON file loaded successfully."
        except Exception as e:
            return False, None, f"Error loading JSON file: {e}"

    def load_jsonl(self):
        try:
            data = []
            with open(self.filepath, 'r', encoding='utf-8') as jsonlfile:
                for line in jsonlfile:
                    data.append(json.loads(line))
            return True, data, "JSONL file loaded successfully."
        except Exception as e:
            return False, None, f"Error loading JSONL file: {e}"

    def load_yaml(self):
        try:
            with open(self.filepath, 'r', encoding='utf-8') as yamlfile:
                data = yaml.safe_load(yamlfile)
            return True, data, "YAML file loaded successfully."
        except Exception as e:
            return False, None, f"Error loading YAML file: {e}"

    def load_ini(self):
        try:
            config = configparser.ConfigParser()
            config.read(self.filepath, encoding='utf-8')
            data = {section: dict(config.items(section)) for section in config.sections()}
            return True, data, "INI file loaded successfully."
        except Exception as e:
            return False, None, f"Error loading INI file: {e}"

    def load_geojson(self):
        return self.load_json()  # GeoJSON is a subset of JSON

    def load_xml(self):
        try:
            tree = ET.parse(self.filepath)
            root = tree.getroot()

            def xml_to_dict(node):
                if len(node) == 0:
                    return node.text if node.text else None
                result = {}
                for child in node:
                    key = child.tag
                    value = xml_to_dict(child)
                    if key in result:
                        if not isinstance(result[key], list):
                            result[key] = [result[key]]
                        result[key].append(value)
                    else:
                        result[key] = value
                return result
            data = xml_to_dict(root)
            return True, data, "XML file loaded successfully."
        except Exception as e:
            return False, None, f"Error loading XML file: {e}"

class DataQuery(Plugin):

    def incompatible_backends(self):
        return [
            'chatgpt-browser',
        ]

    def default_config(self):
        return {
            'agent': {
                'verbose': True,
            },
        }

    def setup(self):
        self.log.info(f"Setting up data query plugin, running with backend: {self.backend.name}")
        self.agent_verbose = self.config.get('plugins.database.agent.verbose')
        self.data_loader = DataLoader(self.config)

    def get_shell_completions(self, _base_shell_completions):
        commands = {}
        commands[util.command_with_leader('data-query')] = {'load': PathCompleter(), 'unload': None}
        return commands

    def load(self, filepath=None):
        self.filepath = filepath
        success, data, user_message = self.data_loader.load(self.filepath)
        if success:
            self.data = data
            # JsonSpec doesn't seem to allow a list at the top level,
            # so wrap lists in a simple 'data' dict.
            if isinstance(self.data, list):
                self.data = {'data': self.data}
            self.log.debug(user_message)
        else:
            self.log.error(user_message)
            return success, self.data, user_message
        json_spec = JsonSpec(dict_=self.data, max_value_length=4000)
        json_toolkit = JsonToolkit(spec=json_spec)
        self.agent = create_json_agent(
            llm=self.make_llm(),
            toolkit=json_toolkit,
            verbose=self.agent_verbose
        )
        return success, data, user_message

    def unload(self):
        self.filepath = None
        self.agent = None
        self.data = None

    def do_data_query(self, arg):
        """
        Send natural language commands to a loaded file of structured data.

        Currently supports the following file types:
            .csv: Comma-separated values
            .geojson: GeoJSON
            .ini: INI
            .json: JSON
            .jsonl: JSONL
            .xlsx: Excel
            .xml: XML
            .yaml: YAML

        Caveats:
            Large datasets will be chunked before sending to the LLM, in this case
            the LLM may not be able to consider the full data set when formulating
            a response.

        Arguments:
            configure: Optional configuration commands, one of:
                load: Load a file of structured data, provide the file path.
                unload: Unload any loaded file.
            prompt: Prompt to interact with the data.

        Examples:
            {COMMAND} load test.csv
            {COMMAND} unload
            {COMMAND} Find users with last name "Smith".
        """
        if not arg:
            return False, arg, "Command is required"
        args = arg.split(maxsplit=1)
        if args[0] == 'load':
            if len(args) > 1:
                success, data, user_message = self.load(args[1])
                return success, data, user_message
            else:
                return False, None, "File path is required"
        if args[0] == 'unload':
            filepath = self.filepath
            self.unload()
            return True, None, f"File {filepath} unloaded"
        if not self.agent:
            return False, None, "No file loaded"
        try:
            result = self.agent.run(arg)
        except ValueError as e:
            return False, arg, e
        return True, arg, result
